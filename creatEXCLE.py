import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '方法步骤'
sheet['A1'] = '使用python创建EXCLE文件'
rows = [['1 创建对象','首先使用openpyxl.Workbook()创建对象'],
        ['2 获取活动工作表','使用对象.active'],
        ['3 命名工作表','工作表.title = '],
        ['4 指定单元格','工作表["单元格位置"] = '],
        ['5 添加行','使用append(列表)'],
        ['6 保存文件','对象.save("文件名")'],
        ['7 读取文件','使用openpyxl.load_workbook("文件名")'],
        ['8 读取工作表','对象["工作表"]'],
        ['9 读取单元格','工作表["单元格位置"].value'],
        ['10','']]
for i in rows:
    sheet.append(i)
# sheet.merge_cells('A10:C15') #start_row=1, start_column=1, end_row=1, end_column=2)
wb.save('使用python创建EXCLE文件.xlsx')

wb = openpyxl.load_workbook('使用python创建EXCLE文件.xlsx')
print(wb.sheetnames)
sheet = wb['方法步骤']
print(sheet['A10'].value)